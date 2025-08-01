import sys
import os
import csv
import json

def parse_csv(csv_path):
    """Extract requirements from a CSV file (Artifact Type='Requirement')."""
    requirements = []
    with open(csv_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            artifact_type = row.get('Artifact Type')
            if artifact_type and artifact_type.strip().lower() == 'requirement':
                req_id = row.get('id')
                text = row.get('Primary Text')
                supplier_comment = row.get('Supplier_1 Comment')
                supplier_status = row.get('Supplier_1 Status')
                if req_id and text:
                    requirements.append({
                        "id": str(req_id),
                        "text": str(text),
                        "supplier_comment": str(supplier_comment) if supplier_comment else "",
                        "supplier_status": str(supplier_status) if supplier_status else ""
                    })
    return requirements

def parse_excel(excel_path):
    """Extract requirements from an Excel file (Artifact Type='Requirement')."""
    import openpyxl
    requirements = []
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        artifact_type = row_dict.get('Artifact Type')
        if artifact_type and str(artifact_type).strip().lower() == 'requirement':
            req_id = row_dict.get('id')
            text = row_dict.get('Primary Text')
            supplier_comment = row_dict.get('Supplier_1 Comment')
            supplier_status = row_dict.get('Supplier_1 Status')
            if req_id and text:
                requirements.append({
                    "id": str(req_id),
                    "text": str(text),
                    "supplier_comment": str(supplier_comment) if supplier_comment else "",
                    "supplier_status": str(supplier_status) if supplier_status else ""
                })
    return requirements

def write_jsonl(requirements, output_path):
    """Write requirements to a JSONL file."""
    with open(output_path, 'w', encoding='utf-8') as f:
        for req in requirements:
            f.write(json.dumps(req, ensure_ascii=False) + '\n')

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python req_to_jsonl.py <input_file> <output_jsonl_file>")
        sys.exit(1)
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    if not os.path.isfile(input_file):
        print(f"Error: Input file '{input_file}' does not exist.")
        sys.exit(1)
    if input_file.lower().endswith('.xlsx'):
        requirements = parse_excel(input_file)
    elif input_file.lower().endswith('.csv'):
        requirements = parse_csv(input_file)
    else:
        print("Error: Unsupported file type. Use .csv or .xlsx")
        sys.exit(1)
    write_jsonl(requirements, output_file)
    print(f"Exported {len(requirements)} requirements to {output_file}.")