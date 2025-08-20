import sys
import json
import torch
import os
import datetime
import shutil
import pandas as pd
from torch.utils.data import DataLoader
from transformers import BertTokenizer, BertForSequenceClassification
from torch.optim import AdamW
import tkinter as tk
from tkinter import messagebox, filedialog, scrolledtext
from tkinter import ttk
from PIL import Image, ImageTk
import csv
import nlpaug.augmenter.word as naw
from sklearn.model_selection import train_test_split

# Import functions from other files
import train_req_model_v2
from train_req_model_v2 import RequirementsDataset
from test_req_model import predict_supplier_status
from req_to_jsonl import parse_csv, parse_excel, write_jsonl

# Set proxy with authentication only if needed
# Uncomment and set your actual proxy if required
# os.environ['HTTP_PROXY'] = 'http://username:password@proxy.company.com:port'
# os.environ['HTTPS_PROXY'] = 'http://username:password@proxy.company.com:port'

def contextual_augment(text, model_name="bert-base-uncased"):
    # Determine device - use GPU if available
    device = "cuda" if torch.cuda.is_available() else "cpu"
    
    # You can switch model_name to "roberta-base" if preferred
    aug = naw.ContextualWordEmbsAug(
        model_path=model_name, action="substitute", device=device
    )
    
    # Print info about which device is being used for augmentation
    print(f"Data augmentation running on: {device}")
    
    return aug.augment(text)

def prepare_datasets(data_path, tokenizer, label_map, val_split=0.2, random_state=42):
    """
    Split data into training and validation sets with stratification.
    """
    # Load all samples
    samples = []
    labels = []
    with open(data_path, encoding='utf-8') as f:
        for line in f:
            item = json.loads(line)
            samples.append(item['text'])
            labels.append(label_map[item['supplier_status'].lower()])
    
    # Stratified split
    X_train, X_val, y_train, y_val = train_test_split(
        samples, labels, 
        test_size=val_split, 
        random_state=random_state,
        stratify=labels
    )
    
    return (
        RequirementsDataset(X_train, y_train, tokenizer),
        RequirementsDataset(X_val, y_val, tokenizer)
    )

class ReqEvalApp:
    def __init__(self, master):
        self.master = master
        master.title("Requirement Evaluation Tool")
        master.geometry("800x700")
        master.configure(bg="#f0f0f0")
        
        # Initialize all attributes first
        self.label_map = {"agreed": 0, "partly agreed": 1, "not agreed": 2}
        self.reverse_label_map = {v: k for k, v in self.label_map.items()}
        self.model = None
        self.tokenizer = None
        self.history = []
        
        # Storage for reinforcement learning
        self.hard_examples = []  # Examples with low confidence
        self.corrections = []    # User-corrected examples
        self.feedback_data = {}  # Stores feedback for each prediction
        self.feedback_file = "feedback_data.json"
        
        # Initialize GUI status
        self.gpu_status = "unknown"
        
        # Check GPU availability on startup
        self.check_gpu_availability()
        
        # Create notebook with tabs
        self.notebook = ttk.Notebook(self.master)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
    def check_gpu_availability(self):
        """Check GPU availability and initialize CUDA"""
        try:
            # Try to initialize CUDA context at startup
            if torch.cuda.is_available():
                # Create a small tensor to initialize CUDA
                _ = torch.zeros(1).cuda()
                torch.cuda.synchronize()
                print(f"CUDA initialized successfully. GPU: {torch.cuda.get_device_name(0)}")
                self.gpu_status = "available"
            else:
                self.gpu_status = "unavailable"
        except Exception as e:
            print(f"CUDA initialization error: {str(e)}")
            self.gpu_status = "error"
        
        # Load existing feedback data if available
        if os.path.exists(self.feedback_file):
            try:
                with open(self.feedback_file, 'r') as f:
                    feedback_data = json.load(f)
                    self.hard_examples = feedback_data.get('hard_examples', [])
                    self.corrections = feedback_data.get('corrections', [])
            except json.JSONDecodeError:
                print("Error reading feedback data file")
            except Exception as e:
                print(f"Unexpected error loading feedback data: {str(e)}")
        
        # Load feedback
        self.load_feedback()
        
        # Create notebook with tabs
        self.notebook = ttk.Notebook(self.master)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create tabs
        self.train_tab = ttk.Frame(self.notebook)
        self.predict_tab = ttk.Frame(self.notebook)
        self.convert_tab = ttk.Frame(self.notebook)
        self.history_tab = ttk.Frame(self.notebook)
        self.model_info_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.train_tab, text="Train Model")
        self.notebook.add(self.predict_tab, text="Predict Status")
        self.notebook.add(self.convert_tab, text="Convert Data")
        self.notebook.add(self.history_tab, text="History")
        self.notebook.add(self.model_info_tab, text="Model Info")
        
        # Setup Training Tab
        self.setup_train_tab()
        
        # Setup Prediction Tab
        self.setup_predict_tab()
        
        # Setup Convert Tab
        self.setup_convert_tab()
        
        # Setup History Tab
        self.setup_history_tab()
        
        # Setup Model Info Tab
        self.setup_model_info_tab()
        
        # Check if model exists and load it
        try:
            if os.path.exists("bert_req_eval_model") and os.path.exists(os.path.join("bert_req_eval_model", "config.json")):
                # Use the safer loading from test_req_model
                from test_req_model import load_model
                self.model, self.tokenizer = load_model()
                self.status_label.config(text="Model loaded successfully")
            else:
                self.status_label.config(text="No model found. Please train a model first.")
        except Exception as e:
            print(f"Error loading model: {str(e)}")  # Print to console for debugging
            self.status_label.config(text=f"Error loading model: {str(e)}")
    def augment_training_data(self):
        data_file = self.data_path.get()
        if not os.path.exists(data_file):
            messagebox.showerror("Error", f"Training data file not found: {data_file}")
            return

        # Create progress window
        progress_window = tk.Toplevel(self.master)
        progress_window.title("Augmenting Data")
        progress_window.geometry("400x200")
        
        # Create progress label and bar
        progress_label = tk.Label(progress_window, text="Initializing augmentation...")
        progress_label.pack(pady=10)
        
        progress_bar = ttk.Progressbar(progress_window, orient=tk.HORIZONTAL, length=300, mode='determinate')
        progress_bar.pack(pady=10)
        
        device_label = tk.Label(progress_window, text="")
        device_label.pack(pady=5)
        
        # Update device info
        device = "GPU (CUDA)" if torch.cuda.is_available() else "CPU"
        device_label.config(text=f"Using: {device}")
        
        progress_window.update()

        try:
            # Count total lines for progress tracking
            total_lines = sum(1 for _ in open(data_file, encoding='utf-8'))
            progress_bar["maximum"] = total_lines
            
            augmented_lines = []
            with open(data_file, encoding='utf-8') as f:
                for i, line in enumerate(f):
                    # Update progress
                    progress_bar["value"] = i + 1
                    progress_label.config(text=f"Augmenting item {i+1} of {total_lines}")
                    progress_window.update()
                    
                    # Process the line
                    item = json.loads(line)
                    orig_text = item.get("text", "")
                    aug_text = contextual_augment(orig_text, model_name="bert-base-uncased")
                    augmented_lines.append(json.dumps({"text": aug_text, "supplier_status": item.get("supplier_status", "")}))
            
            # Update progress for writing phase
            progress_label.config(text="Writing augmented data to file...")
            progress_window.update()
            
            aug_file = os.path.splitext(data_file)[0] + "_augmented.jsonl"
            with open(aug_file, "w", encoding="utf-8") as f:
                with open(data_file, encoding='utf-8') as orig_f:
                    for line in orig_f:
                        f.write(line)
                for line in augmented_lines:
                    f.write(line + "\n")
            
            # Update completion status
            progress_label.config(text="Augmentation complete!")
            progress_bar["value"] = progress_bar["maximum"]
            
            # Add stats to progress window
            stats_text = f"Original samples: {total_lines}\n"
            stats_text += f"New augmented samples: {len(augmented_lines)}\n"
            stats_text += f"Total dataset size: {total_lines + len(augmented_lines)}\n"
            stats_text += f"Device used: {device}"
            
            stats_label = tk.Label(progress_window, text=stats_text, justify="left")
            stats_label.pack(pady=10)
            
            # Add close button to progress window
            ttk.Button(progress_window, text="Close", 
                      command=progress_window.destroy).pack(pady=5)
            
            # Update main application
            self.data_path.set(aug_file)
            self.log_text.insert(tk.END, f"Augmented data created: {aug_file}\n")
            self.log_text.insert(tk.END, f"Used {device} for augmentation\n")
            self.log_text.insert(tk.END, f"Total dataset size: {total_lines + len(augmented_lines)} samples\n")
            self.log_text.see(tk.END)
            
        except Exception as e:
            # Close progress window on error
            progress_window.destroy()
            messagebox.showerror("Augmentation Error", f"Failed to augment data: {str(e)}")
    def show_layer_info(self, info_text):
        """Display information about the selected layer training strategy"""
        self.layer_info_label.config(text=info_text)

    def setup_train_tab(self):
        frame = ttk.LabelFrame(self.train_tab, text="Model Training")
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Data file selection
        file_frame = ttk.Frame(frame)
        file_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(file_frame, text="Training Data:").pack(side="left")
        self.data_path = tk.StringVar(value="requirements_for_llm.jsonl")
        ttk.Entry(file_frame, textvariable=self.data_path, width=40).pack(side="left", padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_data).pack(side="left")
        
        # Training parameters
        param_frame = ttk.LabelFrame(frame, text="Training Parameters")
        param_frame.pack(fill="x", padx=10, pady=5)
        
        # Epochs
        epoch_frame = ttk.Frame(param_frame)
        epoch_frame.pack(fill="x", pady=5)
        ttk.Label(epoch_frame, text="Epochs:").pack(side="left")
        self.epochs = tk.IntVar(value=3)
        ttk.Spinbox(epoch_frame, from_=1, to=10, textvariable=self.epochs, width=5).pack(side="left", padx=5)
        
        # Training parameters frame
        training_params = ttk.LabelFrame(param_frame, text="Training Parameters")
        training_params.pack(fill="x", pady=5, padx=5)
        
        # Batch size
        batch_frame = ttk.Frame(training_params)
        batch_frame.pack(fill="x", pady=5, padx=5)
        ttk.Label(batch_frame, text="Batch Size:").pack(side="left")
        self.batch_size = tk.IntVar(value=8)
        ttk.Spinbox(batch_frame, from_=1, to=32, textvariable=self.batch_size, width=5).pack(side="left", padx=5)
        ttk.Label(batch_frame, text="(smaller = less memory, larger = faster training)").pack(side="left", padx=5)
        
        # Number of epochs
        epoch_frame = ttk.Frame(training_params)
        epoch_frame.pack(fill="x", pady=5, padx=5)
        ttk.Label(epoch_frame, text="Epochs:").pack(side="left")
        self.num_epochs = tk.IntVar(value=3)
        ttk.Spinbox(epoch_frame, from_=1, to=20, textvariable=self.num_epochs, width=5).pack(side="left", padx=5)
        ttk.Label(epoch_frame, text="(more epochs = better learning but risk of overfitting)").pack(side="left", padx=5)
        
        # GPU option with better feedback
        gpu_frame = ttk.LabelFrame(training_params, text="GPU Acceleration")
        gpu_frame.pack(fill="x", pady=5, padx=5)
        
        # GPU use checkbox
        self.use_gpu = tk.BooleanVar(value=True)
        gpu_check_frame = ttk.Frame(gpu_frame)
        gpu_check_frame.pack(fill="x", pady=5, padx=5)
        ttk.Checkbutton(gpu_check_frame, text="Use GPU acceleration", 
                      variable=self.use_gpu).pack(side="left")
        
        # Test GPU button
        ttk.Button(gpu_check_frame, text="Test GPU", 
                 command=lambda: self.run_external_script("src/check_gpu.py")).pack(side="right")
        
        # GPU status frame
        gpu_status_frame = ttk.Frame(gpu_frame)
        gpu_status_frame.pack(fill="x", pady=5, padx=5)
        
        # Visualization option
        visual_frame = ttk.LabelFrame(training_params, text="Visualization")
        visual_frame.pack(fill="x", pady=5, padx=5)
        
        self.visualize_training = tk.BooleanVar(value=False)
        ttk.Checkbutton(visual_frame, text="Show real-time layer visualization during training", 
                      variable=self.visualize_training).pack(side="left", padx=5, pady=5)
        
        # Add two labels for explanation and performance note
        info_frame = ttk.Frame(visual_frame)
        info_frame.pack(fill="x", padx=5)
        
        ttk.Label(info_frame, text="(Shows activations, gradients and training metrics in real-time)",
                font=("", 8, "italic")).pack(side="left", padx=5)
        
        ttk.Label(info_frame, text="Optional: may slightly reduce training speed",
                font=("", 8, "italic"), foreground="#555555").pack(side="right", padx=5)
        
        # GPU detection status
        gpu_info = ""
        if torch.cuda.is_available():
            try:
                gpu_name = torch.cuda.get_device_name(0)
                mem = torch.cuda.get_device_properties(0).total_memory / (1024**3)  # GB
                gpu_info = f"✅ Detected: {gpu_name} ({mem:.1f}GB)"
                self.gpu_status = "available"
            except Exception as e:
                gpu_info = f"⚠️ GPU detected but error: {str(e)}"
                self.gpu_status = "error"
        else:
            gpu_info = "❌ No CUDA GPU detected. Using CPU only."
            self.gpu_status = "unavailable"
        
        # Display GPU status with appropriate styling
        status_label = ttk.Label(gpu_status_frame, text=gpu_info, font=("", 9))
        status_label.pack(side="left", padx=5)
        
        # Add help text if needed
        if self.gpu_status != "available":
            help_text = "See GPU_SETUP.md file for troubleshooting"
            ttk.Label(gpu_status_frame, text=help_text, font=("", 8, "italic"), 
                    foreground="blue").pack(side="right", padx=5)
        
        # Layer architecture info
        layer_frame = ttk.Frame(param_frame)
        layer_frame.pack(fill="x", pady=5)
        ttk.Label(layer_frame, text="Model Architecture:", font=("", 9, "bold")).pack(side="left")
        
        # Info about the fixed layer architecture
        layer_info = (
            "Optimized 6-Layer Training:\n"
            "• Layers 0-5: Frozen (preserves basic language understanding)\n"
            "• Layers 6-8: Trained for technical terminology adaptation\n"
            "• Layers 9-11: Specialized for requirements classification\n"
            "• Classification Head: Fully trained for task optimization"
        )
        
        layer_info_label = ttk.Label(layer_frame, text=layer_info, 
                                   font=("", 8), justify="left",
                                   wraplength=400)
        layer_info_label.pack(side="left", padx=10, pady=5)
        
        # Progress section
        progress_frame = ttk.LabelFrame(frame, text="Training Progress")
        progress_frame.pack(fill="x", padx=10, pady=5)
        
        # Progress bar
        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, length=600, mode='determinate')
        self.progress.pack(fill="x", padx=10, pady=(10,5))
        
        # Status labels frame (using grid for better alignment)
        status_frame = ttk.Frame(progress_frame)
        status_frame.pack(fill="x", padx=10, pady=(0,10))
        
        # Overall progress label
        self.status_label = ttk.Label(status_frame, text="Ready to train", font=("", 9))
        self.status_label.pack(side="left", pady=5)
        
        # Batch progress label
        self.batch_status = ttk.Label(status_frame, text="", font=("", 9))
        self.batch_status.pack(side="right", pady=5)
        
        # Training logs
        log_frame = ttk.LabelFrame(frame, text="Training Logs")
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=10)
        self.log_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Training button
        self.train_button = ttk.Button(frame, text="Train Model", command=self.train_model)
        self.train_button.pack(pady=10)
        
        # Augment button
        self.augment_button = ttk.Button(frame, text="Augment Training Data", command=self.augment_training_data)
        self.augment_button.pack(pady=5)
    
    def setup_predict_tab(self):
        frame = ttk.LabelFrame(self.predict_tab, text="Requirement Evaluation")
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        ttk.Label(frame, text="Enter requirement text:").pack(anchor="w", padx=10, pady=5)
        
        self.req_text = scrolledtext.ScrolledText(frame, height=5)
        self.req_text.pack(fill="both", expand=True, padx=10, pady=5)
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill="x", pady=10)
        
        ttk.Button(button_frame, text="Predict", command=self.predict_requirement).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Clear", command=self.clear_prediction).pack(side="left")
        
        # Feedback frame
        feedback_frame = ttk.LabelFrame(frame, text="Prediction Feedback")
        feedback_frame.pack(fill="x", padx=10, pady=5)
        
        # Correction interface
        correction_frame = ttk.Frame(feedback_frame)
        correction_frame.pack(fill="x", pady=5)
        ttk.Label(correction_frame, text="Correct Status:").pack(side="left", padx=5)
        self.correction_var = tk.StringVar(value="")
        for status in ["agreed", "partly agreed", "not agreed"]:
            ttk.Radiobutton(correction_frame, text=status.capitalize(),
                          variable=self.correction_var, value=status).pack(side="left", padx=5)
        
        # Feedback buttons frame
        feedback_buttons = ttk.Frame(feedback_frame)
        feedback_buttons.pack(fill="x", pady=5)
        
        ttk.Button(feedback_buttons, text="✓ Correct",
                  command=lambda: self.record_feedback(True)).pack(side="left", padx=5)
        ttk.Button(feedback_buttons, text="✗ Incorrect",
                  command=lambda: self.record_feedback(False)).pack(side="left", padx=5)
        ttk.Button(feedback_buttons, text="Submit Correction",
                  command=self.submit_correction).pack(side="left", padx=5)
        
        # Reinforcement learning button
        ttk.Button(feedback_buttons, text="Reinforce Model",
                  command=self.reinforce_model).pack(side="right", padx=5)
        
        result_frame = ttk.LabelFrame(frame, text="Prediction Result")
        result_frame.pack(fill="x", padx=10, pady=10)
        
        self.result_label = ttk.Label(result_frame, text="No prediction yet", font=("Arial", 12))
        self.result_label.pack(pady=10)
        
        # Confidence scores
        self.confidence_frame = ttk.Frame(result_frame)
        self.confidence_frame.pack(fill="x", padx=10, pady=5)
        
        # Add prediction result visualization
        self.confidence_bars = {}
        for status in ["agreed", "partly agreed", "not agreed"]:
            label_frame = ttk.Frame(self.confidence_frame)
            label_frame.pack(fill="x", pady=2)
            
            ttk.Label(label_frame, text=f"{status.capitalize()}: ", width=15).pack(side="left")
            
            self.confidence_bars[status] = ttk.Progressbar(label_frame, orient="horizontal", length=300, mode="determinate")
            self.confidence_bars[status].pack(side="left", padx=5)
            
            self.confidence_bars[status+"_label"] = ttk.Label(label_frame, text="0%", width=5)
            self.confidence_bars[status+"_label"].pack(side="left")
    
    def setup_convert_tab(self):
        frame = ttk.LabelFrame(self.convert_tab, text="Convert CSV/Excel to JSONL")
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Input file selection
        input_frame = ttk.Frame(frame)
        input_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(input_frame, text="Input File:").pack(side="left")
        self.input_path = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.input_path, width=40).pack(side="left", padx=5)
        ttk.Button(input_frame, text="Browse", command=self.browse_input_file).pack(side="left")
        
        # Output file selection
        output_frame = ttk.Frame(frame)
        output_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(output_frame, text="Output JSONL:").pack(side="left")
        self.output_path = tk.StringVar(value="requirements_for_llm.jsonl")
        ttk.Entry(output_frame, textvariable=self.output_path, width=40).pack(side="left", padx=5)
        ttk.Button(output_frame, text="Browse", command=self.browse_output_file).pack(side="left")
        
        # Options frame
        options_frame = ttk.Frame(frame)
        options_frame.pack(fill="x", padx=10, pady=10)
        
        # Add append option
        self.append_mode = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="Append to existing JSONL (instead of overwrite)", 
                      variable=self.append_mode).pack(side="left", padx=5)
        
        # Info label for append mode
        ttk.Label(options_frame, text="Use this when processing multiple files", 
                font=("", 8, "italic")).pack(side="left", padx=5)
        
        # Convert button
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill="x", padx=10, pady=20)
        
        ttk.Button(button_frame, text="Convert to JSONL", command=self.convert_to_jsonl).pack(side="left", padx=10)
        
        # Status label
        self.convert_status = ttk.Label(frame, text="")
        self.convert_status.pack(pady=10)
        
        # Preview area
        preview_frame = ttk.LabelFrame(frame, text="Data Preview")
        preview_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.preview_text = scrolledtext.ScrolledText(preview_frame, height=10)
        self.preview_text.pack(fill="both", expand=True, padx=5, pady=5)
    
    def setup_history_tab(self):
        frame = ttk.Frame(self.history_tab)
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        ttk.Label(frame, text="Prediction History", font=("Arial", 12, "bold")).pack(anchor="w", pady=5)
        
        # Create treeview for history
        columns = ("Requirement", "Prediction")
        self.history_tree = ttk.Treeview(frame, columns=columns, show="headings")
        
        # Set column headings
        self.history_tree.heading("Requirement", text="Requirement")
        self.history_tree.heading("Prediction", text="Prediction")
        
        # Set column widths
        self.history_tree.column("Requirement", width=500)
        self.history_tree.column("Prediction", width=100)
        
        self.history_tree.pack(fill="both", expand=True)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.history_tree.yview)
        scrollbar.pack(side="right", fill="y")
        
        self.history_tree.configure(yscrollcommand=scrollbar.set)
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill="x", pady=10)
        
        ttk.Button(button_frame, text="Export History", command=self.export_history).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Clear History", command=self.clear_history).pack(side="left")
    
    def setup_model_info_tab(self):
        frame = ttk.Frame(self.model_info_tab)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        ttk.Label(frame, text="Model Features & Parameters", font=("Arial", 14, "bold")).pack(anchor="w", pady=10)

        self.model_info_text = scrolledtext.ScrolledText(frame, height=20)
        self.model_info_text.pack(fill="both", expand=True, padx=5, pady=5)

        self.update_model_info()
    
    def update_model_info(self):
        info = []
        
        # Basic Model Information
        info.append("=== Model Configuration ===")
        info.append("Base Model: RoBERTa")
        info.append("Model Variant: roberta-base")
        info.append("Why RoBERTa for Requirements:")
        info.append("  • Better technical text understanding")
        info.append("  • Improved context processing")
        info.append("  • Enhanced handling of formal language")
        info.append("  • Superior performance on domain-specific tasks")
        
        # Safe access to model attribute
        model_status = "Not Loaded"
        if hasattr(self, 'model') and self.model is not None:
            model_status = "Loaded"
            
        info.append(f"Model Status: {model_status}")
        info.append(f"Model Location: bert_req_eval_model")
        info.append("")
        
        # Architecture Details
        info.append("=== Model Architecture ===")
        info.append("Base Architecture: RoBERTa (Technical Requirements Optimized)")
        info.append("Total Layers: 12 transformer layers + classification head")
        info.append("Layer Configuration:")
        info.append("  • Layers 0-5: Frozen (Basic Language Understanding)")
        info.append("    - Syntax and grammar processing")
        info.append("    - Basic semantic relationships")
        info.append("    - Token-level features")
        info.append("  • Layers 6-8: Active (Domain Adaptation)")
        info.append("    - Technical terminology processing")
        info.append("    - Domain-specific patterns")
        info.append("    - Contextual understanding")
        info.append("  • Layers 9-11: Active (Task Specialization)")
        info.append("    - Requirements classification")
        info.append("    - Decision boundaries")
        info.append("    - Category-specific features")
        info.append("  • Classification Head: Active")
        info.append("    - 3-way classification")
        info.append("    - Softmax activation")
        info.append("")
        
        # Training Configuration
        info.append("=== Training Configuration ===")
        info.append(f"Optimizer: AdamW with weight decay")
        info.append(f"Learning Rate: {getattr(self, 'lr', 3e-5)} (Optimized for fine-tuning)")
        info.append(f"Batch Size: {self.batch_size.get() if hasattr(self, 'batch_size') else 'N/A'}")
        info.append(f"Training Epochs: {self.epochs.get() if hasattr(self, 'epochs') else 'N/A'}")
        info.append(f"Loss Function: CrossEntropyLoss (3-class classification)")
        info.append("")
        
        # Data Processing
        info.append("=== Data Processing ===")
        
        # Safe access to tokenizer attribute
        tokenizer_name = "N/A"
        if hasattr(self, 'tokenizer') and self.tokenizer is not None:
            tokenizer_name = type(self.tokenizer).__name__
            
        info.append(f"Tokenizer: {tokenizer_name}")
        info.append("Max Sequence Length: 128 tokens")
        info.append("Text Augmentation: ContextualWordEmbsAug (Technical domain)")
        info.append(f"Training Data: {self.data_path.get() if hasattr(self, 'data_path') else 'N/A'}")
        info.append("")
        
        # Classification Details
        info.append("=== Classification Configuration ===")
        info.append("Categories:")
        for status, idx in self.label_map.items():
            info.append(f"  • {status.capitalize()} (Class {idx})")
        info.append("")
        
        # Reinforcement Learning
        info.append("=== Reinforcement Learning Status ===")
        
        # Safe access to collections
        hard_examples_count = 0
        if hasattr(self, 'hard_examples'):
            hard_examples_count = len(self.hard_examples)
            
        corrections_count = 0
        if hasattr(self, 'corrections'):
            corrections_count = len(self.corrections)
            
        info.append(f"Collected Hard Examples: {hard_examples_count}")
        info.append(f"User Corrections: {corrections_count}")
        
        total_feedback = 0
        if hasattr(self, 'feedback_data'):
            total_feedback = len(self.feedback_data)
            info.append(f"Total Feedback Entries: {total_feedback}")

        self.model_info_text.delete(1.0, tk.END)
        self.model_info_text.insert(tk.END, "\n".join(info))

    def browse_data(self):
        filename = filedialog.askopenfilename(
            title="Select Training Data",
            filetypes=(("JSONL files", "*.jsonl"), ("All files", "*.*"))
        )
        if filename:
            self.data_path.set(filename)
    
    def browse_input_file(self):
        filename = filedialog.askopenfilename(
            title="Select Input File",
            filetypes=(("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*"))
        )
        if filename:
            self.input_path.set(filename)
            self.preview_input_file(filename)
    
    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Save JSONL File",
            defaultextension=".jsonl",
            filetypes=(("JSONL files", "*.jsonl"), ("All files", "*.*"))
        )
        if filename:
            self.output_path.set(filename)
    
    def preview_input_file(self, filename):
        try:
            self.preview_text.delete(1.0, tk.END)
            if filename.lower().endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
                headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                self.preview_text.insert(tk.END, "Headers: " + ", ".join(headers) + "\n\n")
                
                # Preview first 5 rows
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
                    self.preview_text.insert(tk.END, f"Row {i+1}: {str(row)}\n")
            
            elif filename.lower().endswith('.csv'):
                with open(filename, newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    headers = next(reader)
                    self.preview_text.insert(tk.END, "Headers: " + ", ".join(headers) + "\n\n")
                    
                    # Preview first 5 rows
                    for i, row in enumerate(reader):
                        if i >= 5:
                            break
                        self.preview_text.insert(tk.END, f"Row {i+1}: {str(row)}\n")
        
        except Exception as e:
            self.preview_text.insert(tk.END, f"Error previewing file: {str(e)}")
    
    def convert_to_jsonl(self):
        try:
            input_file = self.input_path.get()
            output_file = self.output_path.get()
            
            if not input_file:
                messagebox.showerror("Input Error", "Please select an input file.")
                return
                
            if not os.path.exists(input_file):
                messagebox.showerror("File Error", f"File not found: {input_file}")
                return
            
            self.convert_status.config(text="Converting...")
            self.master.update()
            
            # Use functions from req_to_jsonl.py
            if input_file.lower().endswith('.xlsx'):
                requirements = parse_excel(input_file)
            elif input_file.lower().endswith('.csv'):
                requirements = parse_csv(input_file)
            else:
                messagebox.showerror("File Error", "Unsupported file type. Use .xlsx or .csv")
                return
            
            # Check if we should append
            append_mode = self.append_mode.get()
            
            # Track the number of existing records for status message
            existing_count = 0
            if append_mode and os.path.exists(output_file):
                try:
                    with open(output_file, 'r', encoding='utf-8') as f:
                        existing_count = sum(1 for _ in f)
                except:
                    existing_count = 0
            
            # Write or append to the file
            if append_mode:
                # Custom append function - first check if the file exists
                if os.path.exists(output_file):
                    # Append to existing file
                    with open(output_file, 'a', encoding='utf-8') as f:
                        for req in requirements:
                            f.write(json.dumps(req, ensure_ascii=False) + '\n')
                    status_text = f"Appended {len(requirements)} requirements to {output_file} (Total: {existing_count + len(requirements)})"
                else:
                    # File doesn't exist yet, just write normally
                    write_jsonl(requirements, output_file)
                    status_text = f"Created new file {output_file} with {len(requirements)} requirements"
            else:
                # Normal overwrite mode
                write_jsonl(requirements, output_file)
                status_text = f"Converted {len(requirements)} requirements to {output_file} (Overwrote existing file)"
            
            self.convert_status.config(text=status_text)
            
            # Preview output
            self.preview_text.delete(1.0, tk.END)
            total_lines = 0
            with open(output_file, encoding='utf-8') as f:
                for i, line in enumerate(f):
                    total_lines = i + 1
                    if i < 5:
                        self.preview_text.insert(tk.END, f"{line}\n")
            
            # Add summary at the end
            if total_lines > 5:
                self.preview_text.insert(tk.END, f"\n... and {total_lines - 5} more entries (total: {total_lines})")
                    
            messagebox.showinfo("Conversion Complete", f"Converted {len(requirements)} requirements to {output_file}")
            
        except Exception as e:
            self.convert_status.config(text=f"Error: {str(e)}")
            messagebox.showerror("Conversion Error", f"Failed to convert file: {str(e)}")
    
    def train_model(self):
        try:
            data_file = self.data_path.get()
            if not os.path.exists(data_file):
                messagebox.showerror("Error", f"Training data file not found: {data_file}")
                return
                
            # Reset status message and progress bar at the beginning of training
            self.status_label.config(text="Training in progress...")
            self.log_text.delete(1.0, tk.END)
            self.progress["value"] = 0
            self.master.update_idletasks()
            
            # Track training progress
            self.current_epoch = 0
            self.total_epochs = 3  # Default value from train_req_model_v2
            
            # Define callbacks for training progress
            def on_log(msg):
                self.log_text.insert(tk.END, msg + "\n")
                self.log_text.see(tk.END)
                self.master.update_idletasks()
            
            def on_progress(step, total):
                # Calculate overall progress including epochs
                overall_progress = (self.current_epoch * 100 + (step / total) * 100) / self.total_epochs
                self.progress["value"] = overall_progress
                
                # Update status labels
                self.status_label.config(
                    text=f"Training Progress - Epoch {self.current_epoch + 1}/{self.total_epochs}"
                )
                self.batch_status.config(
                    text=f"Batch: {step}/{total}"
                )
                self.master.update_idletasks()
            
            def on_epoch_end(epoch, metrics):
                self.current_epoch = metrics['epoch']
                msg = (f"Epoch {metrics['epoch']}/{self.total_epochs}: "
                      f"Train Loss = {metrics['train_loss']:.4f}, "
                      f"Train Acc = {metrics['train_acc']:.4f}, "
                      f"Val Loss = {metrics['val_loss']:.4f}, "
                      f"Val Acc = {metrics['val_acc']:.4f}\n")
                self.log_text.insert(tk.END, msg)
                self.log_text.see(tk.END)
                # Update progress bar to show completed epoch
                self.progress["value"] = (metrics['epoch'] * 100) / self.total_epochs
                self.master.update_idletasks()
            
            callbacks = {
                'on_log': on_log,
                'on_progress': on_progress,
                'on_epoch_end': on_epoch_end
            }
            
            # Get training parameters from GUI
            batch_size = self.batch_size.get()
            epochs = self.num_epochs.get()
            
            # Update total_epochs for progress tracking
            self.total_epochs = epochs
            
            try:
                # Clear any existing model from memory
                if hasattr(self, 'model'):
                    del self.model
                if hasattr(self, 'tokenizer'):
                    del self.tokenizer
                
                import gc
                gc.collect()  # Force garbage collection
                
                # Call the training function with callbacks and user parameters
                use_gpu = self.use_gpu.get() if hasattr(self, 'use_gpu') else True
                visualize = self.visualize_training.get() if hasattr(self, 'visualize_training') else False
                
                self.model, self.tokenizer, history = train_req_model_v2.train_model(
                    data_path=data_file,
                    model_dir="bert_req_eval_model",
                    use_roberta=True,
                    batch_size=batch_size,
                    epochs=epochs,
                    callbacks=callbacks,
                    use_gpu=use_gpu,
                    visualize=visualize,
                    parent_window=self.master
                )
                
                # Ensure progress bar shows 100% completion
                self.progress["value"] = 100
                
                # Update status with final metrics
                final_metrics = {
                    'train_loss': history['train_loss'][-1],
                    'train_acc': history['train_acc'][-1],
                    'val_loss': history['val_loss'][-1],
                    'val_acc': history['val_acc'][-1]
                }
                
                completion_message = (
                    f"Training completed successfully!\n"
                    f"Final metrics:\n"
                    f"Training Accuracy: {final_metrics['train_acc']:.2%}\n"
                    f"Validation Accuracy: {final_metrics['val_acc']:.2%}"
                )
                
                # Display completion message
                self.status_label.config(text="✓ Training completed successfully!")
                self.log_text.insert(tk.END, "\n" + completion_message + "\n")
                self.log_text.see(tk.END)
                messagebox.showinfo("Training Complete", completion_message)
                
                # Show training plot if available
                if os.path.exists("bert_req_eval_model/training_history.png"):
                    img = Image.open("bert_req_eval_model/training_history.png")
                    img = img.resize((600, 300), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    
                    # Create a new window for the plot
                    plot_window = tk.Toplevel(self.master)
                    plot_window.title("Training History")
                    
                    # Add the image to a label
                    label = ttk.Label(plot_window, image=photo)
                    label.image = photo  # Keep a reference to prevent garbage collection
                    label.pack()
                
            except Exception as e:
                error_msg = f"Error during model saving: {str(e)}\n\nYour training progress is not lost - the model will be loaded from the last successful save."
                self.log_text.insert(tk.END, f"\nError: {str(e)}\n")
                messagebox.showerror("Training Error", error_msg)
                
                # Try to recover the model if possible
                if os.path.exists("bert_req_eval_model.bak"):
                    try:
                        if os.path.exists("bert_req_eval_model"):
                            shutil.rmtree("bert_req_eval_model")
                        os.rename("bert_req_eval_model.bak", "bert_req_eval_model")
                        self.log_text.insert(tk.END, "Successfully restored model from backup.\n")
                    except Exception as restore_error:
                        self.log_text.insert(tk.END, f"Error restoring backup: {str(restore_error)}\n")
            
        except Exception as e:
            self.log_text.insert(tk.END, f"Error during training: {str(e)}\n")
            self.log_text.see(tk.END)
            self.status_label.config(text="Training error")
            messagebox.showerror("Error", f"Training failed: {str(e)}")
    
    def predict_requirement(self):
        try:
            req_text = self.req_text.get("1.0", tk.END).strip()
            if not req_text:
                messagebox.showwarning("Input Error", "Please enter a requirement text.")
                return
            
            # Check if model exists
            if not os.path.exists("bert_req_eval_model"):
                messagebox.showerror("Model Error", "Model not found. Please train a model first.")
                return
            
            # Use the predict_with_confidence function from test_req_model.py
            from test_req_model import predict_with_confidence
            
            # Get prediction and confidence scores
            result, confidence_scores = predict_with_confidence(req_text)
            
            self.result_label.config(text=f"Predicted: {result.upper()}")
            
            # Store last prediction and confidence for feedback
            self.last_prediction = result
            self.last_confidence = max(confidence_scores.values())
            
            # Update confidence bars
            for status in ["agreed", "partly agreed", "not agreed"]:
                prob_value = confidence_scores[status]
                self.confidence_bars[status]["value"] = prob_value
                self.confidence_bars[status+"_label"].config(text=f"{prob_value:.1f}%")
                
            # Check if this is a low confidence prediction
            if self.last_confidence < 0.8:
                self.result_label.config(text=f"Predicted: {result.upper()} (Low Confidence)")
                self.hard_examples.append({
                    'text': req_text,
                    'prediction': result,
                    'confidence': self.last_confidence
                })
            
            # Add to history
            self.history.append((req_text, result))
            self.history_tree.insert("", "end", values=(req_text[:80] + "..." if len(req_text) > 80 else req_text, result))
            
        except Exception as e:
            messagebox.showerror("Error", f"Prediction error: {str(e)}")
            # Try to reload the model
            try:
                from test_req_model import load_model
                self.model, self.tokenizer = load_model()
                messagebox.showinfo("Model Reloaded", "The model was reloaded. Please try prediction again.")
            except Exception as reload_error:
                messagebox.showerror("Model Error", f"Could not load model: {str(reload_error)}")
    
    def clear_prediction(self):
        self.req_text.delete("1.0", tk.END)
        self.result_label.config(text="No prediction yet")
        for status in ["agreed", "partly agreed", "not agreed"]:
            self.confidence_bars[status]["value"] = 0
            self.confidence_bars[status+"_label"].config(text="0%")
    
    def export_history(self):
        if not self.history:
            messagebox.showinfo("Export", "No prediction history to export.")
            return
            
        filename = filedialog.asksaveasfilename(
            title="Export History",
            defaultextension=".csv",
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*"))
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write("Requirement,Prediction\n")
                    for req, pred in self.history:
                        # Escape quotes and remove newlines
                        req = req.replace('"', '""').replace('\n', ' ').replace('\r', '')
                        f.write(f'"{req}",{pred}\n')
                messagebox.showinfo("Export", f"History exported to {filename}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export history: {str(e)}")
    
    def clear_history(self):
        if messagebox.askyesno("Clear History", "Are you sure you want to clear the prediction history?"):
            self.history = []
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)

    def load_feedback(self):
        """Load saved feedback and corrections"""
        try:
            if os.path.exists(self.feedback_file):
                with open(self.feedback_file, 'r') as f:
                    data = json.load(f)
                    self.feedback_data = data.get('feedback', {})
                    self.hard_examples = data.get('hard_examples', [])
                    self.corrections = data.get('corrections', [])
        except Exception as e:
            print(f"Error loading feedback data: {e}")

    def save_feedback(self):
        """Save feedback and corrections"""
        try:
            data = {
                'feedback': self.feedback_data,
                'hard_examples': self.hard_examples,
                'corrections': self.corrections
            }
            with open(self.feedback_file, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            print(f"Error saving feedback data: {e}")

    def record_feedback(self, is_correct):
        """Record user feedback on prediction"""
        current_text = self.req_text.get("1.0", tk.END).strip()
        if not current_text or not hasattr(self, 'last_prediction'):
            messagebox.showwarning("Feedback Error", "No prediction to provide feedback for.")
            return

        # Store feedback
        self.feedback_data[current_text] = {
            'prediction': self.last_prediction,
            'is_correct': is_correct,
            'confidence': self.last_confidence,
            'timestamp': str(datetime.datetime.now())
        }

        # If incorrect or low confidence, add to hard examples
        if not is_correct or self.last_confidence < 0.8:
            self.hard_examples.append({
                'text': current_text,
                'prediction': self.last_prediction,
                'confidence': self.last_confidence
            })

        self.save_feedback()
        messagebox.showinfo("Feedback Recorded", 
                          "Thank you for your feedback! This will help improve the model.")

    def submit_correction(self):
        """Submit a correction for the current prediction"""
        current_text = self.req_text.get("1.0", tk.END).strip()
        correction = self.correction_var.get()

        if not current_text or not correction:
            messagebox.showwarning("Correction Error", 
                                 "Please enter requirement text and select the correct status.")
            return

        # Store correction
        self.corrections.append({
            'text': current_text,
            'original_prediction': getattr(self, 'last_prediction', None),
            'corrected_status': correction,
            'timestamp': str(datetime.datetime.now())
        })

        self.save_feedback()
        messagebox.showinfo("Correction Recorded", 
                          "Correction saved! This will be used in the next reinforcement training.")

    def reinforce_model(self):
        """Start reinforcement learning using collected feedback"""
        if not self.corrections and not self.hard_examples:
            messagebox.showinfo("Reinforcement Learning", 
                              "No corrections or hard examples collected yet. "
                              "Please provide feedback and corrections first.")
            return

        if messagebox.askyesno("Reinforce Model", 
                             f"Start reinforcement learning with:\n"
                             f"- {len(self.corrections)} corrections\n"
                             f"- {len(self.hard_examples)} hard examples\n"
                             f"This may take a while. Continue?"):
            # Start reinforcement learning
            self.train_model(reinforce_learning=True)
    
    def run_external_script(self, script_path):
        """Run an external Python script and display results in a popup window"""
        try:
            import subprocess
            import sys
            import os
            
            # Use virtual environment Python if available, otherwise use current interpreter
            venv_path = os.path.join(os.getcwd(), ".venv", "Scripts", "python.exe")
            if os.path.exists(venv_path):
                python_exe = venv_path
            else:
                python_exe = sys.executable
            
            # Run the script and capture output
            result = subprocess.run([python_exe, script_path], 
                                  capture_output=True, text=True, check=False)
            
            # Create popup window to display results
            result_window = tk.Toplevel(self.master)
            result_window.title("GPU Test Results")
            result_window.geometry("700x500")
            
            # Add scrolled text widget
            text_widget = scrolledtext.ScrolledText(result_window, wrap=tk.WORD)
            text_widget.pack(expand=True, fill="both", padx=10, pady=10)
            
            # Insert information about which Python interpreter was used
            text_widget.insert(tk.END, f"Using Python interpreter: {python_exe}\n\n")
            
            # Insert the result
            if result.stdout:
                text_widget.insert(tk.END, result.stdout)
            
            # If there's an error, show that too
            if result.stderr:
                text_widget.insert(tk.END, "\n\nERRORS:\n", "error")
                text_widget.insert(tk.END, result.stderr)
                text_widget.tag_configure("error", foreground="red")
            
            # Add close button
            ttk.Button(result_window, text="Close", 
                     command=result_window.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to run script: {str(e)}")
    
    def reset_gui(self):
        # Clear status label
        self.status_label.config(text="")
        # Clear log text
        self.log_text.delete(1.0, tk.END)
        # Clear requirement text input
        if hasattr(self, 'req_text'):
            self.req_text.delete("1.0", tk.END)
        # Clear prediction result label
        if hasattr(self, 'result_label'):
            self.result_label.config(text="")
        # Reset confidence bars
        if hasattr(self, 'confidence_bars'):
            for status in self.confidence_bars:
                self.confidence_bars[status]["value"] = 0
                self.confidence_bars[status+"_label"].config(text="0.0%")
        # Clear history
        if hasattr(self, 'history_tree'):
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)
        # Clear convert status
        if hasattr(self, 'convert_status'):
            self.convert_status.config(text="")
        # Optionally reset other fields as needed

if __name__ == "__main__":
    root = tk.Tk()
    app = ReqEvalApp(root)
    root.mainloop()